# using local copy of yfinance.  I had to modify it to pull additional info from yahoo finance
import yfinance as yf
import openpyxl
import argparse
from bs4 import BeautifulSoup

class Fund:
    def __init__(self):
        self.name                    = ''
        self.fundType                = ''
        self.sym                     = ''
        self.quote                   = 0
        self.qty                     = 0
        self.beta_5y                 = 0
        self.expenseRatio            = 0
        self.morningstarRating       = ''
        self.morningstarRiskRating   = ''
        self.yield_12m               = 0
        self.return_1y               = 0
        self.return_3y               = 0
        self.return_5y               = 0

"""
Read the fund names, symbols, and quantities from the web pages saved 
from Personal Capital.  We will augment this with additional data from 
Yahoo Finance.
"""
def readTable(table):
    funds = {}
    symbols = table.find_all(class_='u-text-bold qa-ticker')
    qtys = table.find_all(class_='table__column table__column--right pc-holdings-grid-cell--holding-shares qa-holding-shares')
    for sym, qty in zip(symbols, qtys):
        f = Fund()
        f.name  = sym['title']
        f.sym   = sym.get_text()
        f.qty   = float(qty.get_text())
        if f.name == 'Cash':
            f.sym = 'CASH'
        funds[f.sym] = f
    return funds

"""
Yahoo Finance sends us the number of Morningstar stars, but rather than
a number, we'd like to show the stars.  For example, rather than displaying
the number 4, show ****.
"""
def buildStarString(i):
    ret = ''
    if i and (i > 0):
        for n in range(i):
            ret = ret + '★'

    return ret

"""
Convert the Morningstar risk ratings from numbers to 
human readable strings.
"""
def getRiskRating(i):
    ret = ''
    if i == 1:
        ret = 'Low'
    elif i == 2:
        ret = 'Below Avg'
    elif i == 3:
        ret = 'Avg'
    elif i == 4:
        ret = 'Above Avg'
    elif i == 5:
        ret = 'High'
    return ret

"""
Iterate through the securities we read from the Personal Capital web page,
and using the symbol, lookup additional info about each security on Yahoo Finance.
"""
def lookupData(symbols):
    print('Yahoo data')
    results = {}
    for sym, f in symbols.items():
        ticker = yf.Ticker(sym)
        if sym == 'CASH':
            f.quote = 1
            f.fundType = 'CASH'
        else:
            hist = ticker.history()
            f.quote = hist['Close'][-1]
            if 'categoryName' in ticker.info:
                f.fundType = ticker.info['categoryName']
            if 'beta3Year' in ticker.info:
                f.beta_5y = ticker.info['beta3Year']  # the element name is 3 year, but the web page says it's 5 years
            if 'annualReportExpenseRatio' in ticker.info:
                f.expenseRatio = ticker.info['annualReportExpenseRatio']
            if 'morningStarOverallRating' in ticker.info:
                f.morningstarRating = ticker.info['morningStarOverallRating']
            if 'morningStarRiskRating' in ticker.info:
                f.morningstarRiskRating = ticker.info['morningStarRiskRating']
            if 'yield' in ticker.info:
                f.yield_12m = ticker.info['yield']
            if 'trailingReturns' in ticker.info:    # monthly pretax returns annualized
                if 'oneYear' in ticker.info['trailingReturns']:
                    f.return_1y = ticker.info['trailingReturns']['oneYear']
                if 'threeYear' in ticker.info['trailingReturns']:
                    f.return_3y = ticker.info['trailingReturns']['threeYear']
                if 'fiveYear' in ticker.info['trailingReturns']:
                    f.return_5y = ticker.info['trailingReturns']['fiveYear']
        results[sym] = f
        print(sym, f.quote)

    return results


"""
Write the header and each fund to the specified spreadsheet.
"""
def writeHeader(sheet):
    sheet.cell(row=1, column=1).value = "Security"
    sheet.cell(row=1, column=2).value = "Symbol"
    sheet.cell(row=1, column=3).value = "Type"
    sheet.cell(row=1, column=4).value = "Qty"
    sheet.cell(row=1, column=5).value = "Quote"
    sheet.cell(row=1, column=6).value = "Rating"
    sheet.cell(row=1, column=7).value = "Risk"
    sheet.cell(row=1, column=8).value = "1 yr return"
    sheet.cell(row=1, column=9).value = "3 yr return"
    sheet.cell(row=1, column=10).value = "5 yr return"
    sheet.cell(row=1, column=11).value = "Beta"
    sheet.cell(row=1, column=12).value = "Expense Ratio"
    sheet.cell(row=1, column=13).value = "Yield"
    sheet.cell(row=1, column=14).value = "Est Annual Income"
    sheet.cell(row=1, column=15).value = "Balance"

def WriteRow(summarySheet, r, fund):
    summarySheet.cell(row=nRow, column=1).value = fund.name
    summarySheet.cell(row=nRow, column=2).value = sym
    summarySheet.cell(row=nRow, column=3).value = fund.fundType
    summarySheet.cell(row=nRow, column=4).value = fund.qty
    summarySheet.cell(row=nRow, column=5).value = fund.quote
    summarySheet.cell(row=nRow, column=6).value = buildStarString(fund.morningstarRating)
    summarySheet.cell(row=nRow, column=7).value = getRiskRating(fund.morningstarRiskRating)
    summarySheet.cell(row=nRow, column=8).value = fund.return_1y
    summarySheet.cell(row=nRow, column=9).value = fund.return_3y
    summarySheet.cell(row=nRow, column=10).value = fund.return_5y
    summarySheet.cell(row=nRow, column=11).value = fund.beta_5y
    summarySheet.cell(row=nRow, column=12).value = fund.expenseRatio
    summarySheet.cell(row=nRow, column=13).value = fund.yield_12m
    summarySheet.cell(row=nRow, column=14).value = (fund.qty * fund.quote) * fund.yield_12m
    summarySheet.cell(row=nRow, column=15).value = fund.qty * fund.quote


if __name__ == '__main__':
    print('exportPortfolio v0.1')
    parser = argparse.ArgumentParser()
    parser.add_argument("-i", "--input", help="Input file with portfolio saved from Personal Capital", required=True)
    parser.add_argument("-o", "--output", help="output file name (.xlsx)", default="output.xlsx")
    parser.add_argument("-y", '--yahoo', dest='yahoo', action='store_true')
    parser.add_argument('--no-yahoo', dest='yahoo', action='store_false')
    parser.set_defaults(yahoo=False)
    args = parser.parse_args()

    with open(args.input, 'r') as file:
        data = file.read().replace('\n', '')

    soup = BeautifulSoup(data, 'html.parser')
    table = soup.find('div', class_='table__body')

    funds = readTable(table)

    if args.yahoo:
        funds = lookupData(funds)

    wb = openpyxl.Workbook()
    summarySheet = wb.active
    summarySheet.title = 'Portfolio Summary'
    writeHeader(summarySheet)
    nRow = 2 # skip the header
    for sym, f in funds.items():
        WriteRow(summarySheet, nRow, f)
        nRow += 1

    wb.save(args.output)
    print('done')